import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Side

# ===== 資料重排函式 =====
def reshape_df(df, rows_per_page=50):
    """把資料重排成每行兩筆，並在每頁插入標頭"""
    group_size = len(df.columns)
    rows = []
    header = df.columns.tolist() + df.columns.tolist()

    for i in range(0, len(df), 2):
        row = []
        row.extend(df.iloc[i].tolist())
        if i + 1 < len(df):
            row.extend(df.iloc[i + 1].tolist())
        else:
            row.extend([""] * group_size)
        rows.append(row)

    # 插入分頁標頭
    final_rows = []
    for i, row in enumerate(rows):
        if i % rows_per_page == 0:
            final_rows.append(header)
        final_rows.append(row)

    return pd.DataFrame(final_rows)

# ===== 套用格式 (粗體 & 框線) =====
def apply_formatting(filepath):
    wb = load_workbook(filepath)
    ws = wb.active

    # 框線樣式
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for row in ws.iter_rows():
        for cell in row:
            # 加框線
            cell.border = thin_border
            # 標頭列粗體 (判斷是否為第一列，或是插入的 header)
            if cell.row == 1 or (ws.cell(row=cell.row, column=1).value == ws.cell(row=1, column=1).value):
                cell.font = Font(bold=True)

    wb.save(filepath)

# ===== GUI 主程式 =====
class ExcelApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Excel 資料重排工具")
        self.root.geometry("600x250")

        # 上方 Frame (選擇 Excel 檔案)
        frame_top = tk.Frame(root, pady=10)
        frame_top.pack(fill="x")

        self.file_label = tk.Label(frame_top, text="尚未選擇檔案", width=50, anchor="w")
        self.file_label.pack(side="left", padx=5)

        btn_choose = tk.Button(frame_top, text="選擇 Excel", command=self.choose_file)
        btn_choose.pack(side="left", padx=5)

        # 中間 Frame (選擇工作表)
        frame_middle = tk.Frame(root, pady=10)
        frame_middle.pack(fill="x")

        tk.Label(frame_middle, text="選擇工作表：").pack(side="left", padx=5)
        self.sheet_combo = ttk.Combobox(frame_middle, state="readonly", width=30)
        self.sheet_combo.pack(side="left", padx=5)

        # 下方 Frame (產生 Excel)
        frame_bottom = tk.Frame(root, pady=20)
        frame_bottom.pack(fill="x")

        btn_generate = tk.Button(frame_bottom, text="產生新 Excel", command=self.generate_excel)
        btn_generate.pack()

        # 狀態
        self.filepath = None
        self.sheets = []

    def choose_file(self):
        filepath = filedialog.askopenfilename(
            title="選擇 Excel 檔案",
            filetypes=[("Excel 檔案", "*.xlsx *.xls")]
        )
        if filepath:
            self.filepath = filepath
            self.file_label.config(text=os.path.basename(filepath))

            try:
                xls = pd.ExcelFile(filepath)
                self.sheets = xls.sheet_names
                self.sheet_combo["values"] = self.sheets
                if self.sheets:
                    self.sheet_combo.current(0)
            except Exception as e:
                messagebox.showerror("錯誤", f"無法讀取工作表：\n{e}")

    def generate_excel(self):
        if not self.filepath:
            messagebox.showwarning("提醒", "請先選擇 Excel 檔案")
            return

        if not self.sheet_combo.get():
            messagebox.showwarning("提醒", "請先選擇工作表")
            return

        try:
            df = pd.read_excel(self.filepath, sheet_name=self.sheet_combo.get())
            reshaped_df = reshape_df(df, rows_per_page=50)  # 預設每頁 50 行

            save_path = filedialog.asksaveasfilename(
                title="另存新檔",
                defaultextension=".xlsx",
                filetypes=[("Excel 檔案", "*.xlsx")]
            )
            if save_path:
                reshaped_df.to_excel(save_path, index=False, header=False)
                apply_formatting(save_path)  # 套用粗體 & 框線
                messagebox.showinfo("完成", f"已輸出新 Excel：\n{save_path}")
        except Exception as e:
            messagebox.showerror("錯誤", f"處理失敗：\n{e}")


if __name__ == "__main__":
    root = tk.Tk()
    app = ExcelApp(root)
    root.mainloop()
